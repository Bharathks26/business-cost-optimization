import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.recommendations import generate_cost_recommendations

def format_worksheet(ws, title):
    """Apply professional formatting to worksheet"""
    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting to first row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Apply borders to all cells with data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
def sanitize_dataframe_for_excel(df):
    df = df.copy()
    for col in df.columns:
        if str(df[col].dtype).startswith("period"):
            df[col] = df[col].astype(str)
        elif "datetime" in str(df[col].dtype):
            df[col] = df[col].astype(str)
        elif df[col].dtype == "object":
            df[col] = df[col].astype(str)
    return df

def generate_business_report(df, kpis, anomalies, forecast_data, username='user'):
    os.makedirs('reports', exist_ok=True)
    
    wb = Workbook()
    
    # Sheet 1: KPI Analysis Results
    ws_kpis = wb.active
    ws_kpis.title = "KPI Analysis"
    ws_kpis.append(['KPI Type', 'Metric', 'Value'])
    
    if kpis.get('monthly_cost'):
        for month, cost in kpis['monthly_cost'].items():
            ws_kpis.append(['Monthly Cost', str(month), f'₹{cost:,.0f}'])
    
    if kpis.get('department_cost'):
        for dept, cost in kpis['department_cost'].items():
            ws_kpis.append(['Department Cost', dept, f'₹{cost:,.0f}'])
    
    if kpis.get('cost_growth_rate'):
        for month, rate in kpis['cost_growth_rate'].items():
            ws_kpis.append(['Growth Rate', str(month), f'{rate*100:.1f}%'])
    
    # Sheet 2: Anomaly Detection Results
    ws_anomalies = wb.create_sheet("Anomaly Detection")
    ws_anomalies.append(['Department', 'Vendor', 'Amount', 'Date', 'Anomaly Score', 'Why Flagged?'])
    
    if not anomalies.empty:
        anomaly_records = anomalies[anomalies['anomaly_flag'] == 1]
        for _, row in anomaly_records.iterrows():
            explanation = row.get('explanation', 'Statistical deviation detected')
            ws_anomalies.append([row['department'], row['vendor'], f'₹{row["amount"]:,.0f}', str(row['date']), f'{row["anomaly_score"]:.3f}', explanation])
    else:
        ws_anomalies.append(['No anomalies detected', 'All expenses follow normal patterns', '', '', '', ''])
    
    # Sheet 3: Cost Optimization Recommendations
    ws_recommendations = wb.create_sheet("Recommendations")
    ws_recommendations.append(['Recommendation', 'Reason / Justification'])
    
    if not df.empty:
        recommendations = generate_cost_recommendations(df, kpis, anomalies.to_dict('records') if not anomalies.empty else [], forecast_data)
        for rec in recommendations:
            if isinstance(rec, dict):
                ws_recommendations.append([rec.get('recommendation', 'Optimization suggestion'), rec.get('reason', 'Based on data analysis')])
            else:
                ws_recommendations.append([str(rec), 'Statistical analysis recommendation'])
    else:
        ws_recommendations.append(['No recommendations available', 'Upload expense data to generate insights'])
    
    # Sheet 4: Expense Forecasting
    ws_forecast = wb.create_sheet("Forecasting")
    ws_forecast.append(['Type', 'Month', 'Predicted Amount'])
    
    if forecast_data.get('historical'):
        for i, month in enumerate(forecast_data['historical']['months']):
            amount = forecast_data['historical']['amounts'][i] if i < len(forecast_data['historical']['amounts']) else 0
            ws_forecast.append(['Historical Trend', str(month), f'₹{amount:,.0f}'])
    
    if forecast_data.get('forecasted'):
        for i, month in enumerate(forecast_data['forecasted']['months']):
            amount = forecast_data['forecasted']['amounts'][i] if i < len(forecast_data['forecasted']['amounts']) else 0
            ws_forecast.append(['Future Prediction', str(month), f'₹{amount:,.0f}'])
    
    # Apply formatting to all sheets
    format_worksheet(ws_kpis, "KPI Analysis")
    format_worksheet(ws_anomalies, "Anomaly Detection")
    format_worksheet(ws_recommendations, "Recommendations")
    format_worksheet(ws_forecast, "Forecasting")
    
    # Generate user-specific filename
    filename = os.path.abspath(f'reports/{username}_business_cost_report.xlsx')
    wb.save(filename)
    
    return filename